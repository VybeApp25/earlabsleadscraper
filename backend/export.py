import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime


def generate_csv(businesses: list[dict], analyses: dict, owners: dict) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)

    headers = [
        "Business Name", "Status", "Niche", "City", "Address",
        "Phone", "Emails", "Website",
        "Google Rating", "Google Reviews",
        "Owner Name", "Owner Title", "Personal Phone", "Personal Email", "Owner LinkedIn",
        "Social Proof Score", "Opportunity Score", "Website Outdated",
        "Monthly Traffic", "Domain Age (Years)",
        "Pain Points", "Website Summary",
        "Instagram", "Facebook", "LinkedIn", "TikTok",
        "Found At",
    ]
    writer.writerow(headers)

    for biz in businesses:
        bid = biz["id"]
        analysis = analyses.get(bid, {})
        owner = owners.get(bid, {})
        social = biz.get("social_links", {})

        writer.writerow([
            biz.get("name", ""),
            biz.get("status", "new"),
            biz.get("niche", ""),
            biz.get("city", ""),
            biz.get("address", ""),
            biz.get("phone", ""),
            "; ".join(biz.get("emails", [])),
            biz.get("website", ""),
            biz.get("google_rating", ""),
            biz.get("google_review_count", ""),
            owner.get("owner_name", ""),
            owner.get("owner_title", ""),
            owner.get("personal_phone", ""),
            owner.get("personal_email", ""),
            owner.get("linkedin_url", ""),
            analysis.get("social_proof_score", ""),
            analysis.get("opportunity_score", ""),
            "Yes" if analysis.get("website_is_outdated") else "No" if analysis else "",
            analysis.get("traffic_monthly", ""),
            analysis.get("website_age_years", ""),
            "; ".join(analysis.get("pain_points", [])),
            analysis.get("website_summary", ""),
            social.get("instagram", ""),
            social.get("facebook", ""),
            social.get("linkedin", ""),
            social.get("tiktok", ""),
            biz.get("found_at", ""),
        ])

    return output.getvalue().encode("utf-8")


def generate_excel(businesses: list[dict], analyses: dict, owners: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "EAR Labs Leads"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1a2033")
    header_font = Font(bold=True, color="4f6ef7", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        ("Business Name", 28), ("Status", 14), ("Niche", 16), ("City", 16),
        ("Phone", 18), ("Emails", 30), ("Website", 30),
        ("Google Rating", 14), ("Reviews", 10),
        ("Owner Name", 22), ("Owner Title", 18), ("Personal Phone", 18),
        ("Personal Email", 28), ("Owner LinkedIn", 30),
        ("Social Score", 13), ("Oppty Score", 13), ("Site Outdated", 14),
        ("Monthly Traffic", 16), ("Domain Age", 12),
        ("Pain Points", 40), ("Instagram", 25), ("Facebook", 25),
        ("Found At", 20),
    ]

    ws.row_dimensions[1].height = 30
    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Status colors
    status_colors = {
        "new": "2d3654",
        "researched": "1e3a5f",
        "contacted": "1e4a3a",
        "replied": "2d4a1e",
        "proposal": "3a3a1e",
        "won": "1a4a1a",
        "lost": "4a1a1a",
    }

    for row_idx, biz in enumerate(businesses, 2):
        bid = biz["id"]
        analysis = analyses.get(bid, {})
        owner = owners.get(bid, {})
        social = biz.get("social_links", {})

        row_data = [
            biz.get("name", ""),
            biz.get("status", "new").upper(),
            biz.get("niche", ""),
            biz.get("city", ""),
            biz.get("phone", ""),
            "; ".join(biz.get("emails", [])),
            biz.get("website", ""),
            biz.get("google_rating", ""),
            biz.get("google_review_count", ""),
            owner.get("owner_name", ""),
            owner.get("owner_title", ""),
            owner.get("personal_phone", ""),
            owner.get("personal_email", ""),
            owner.get("linkedin_url", ""),
            analysis.get("social_proof_score", ""),
            analysis.get("opportunity_score", ""),
            "Yes" if analysis.get("website_is_outdated") else "No" if analysis else "",
            analysis.get("traffic_monthly", ""),
            analysis.get("website_age_years", ""),
            "; ".join(analysis.get("pain_points", [])),
            social.get("instagram", ""),
            social.get("facebook", ""),
            biz.get("found_at", ""),
        ]

        status = biz.get("status", "new")
        row_fill = PatternFill("solid", fgColor=status_colors.get(status, "111827"))

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = Font(color="e2e8f0", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.row_dimensions[row_idx].height = 18

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    bg = PatternFill("solid", fgColor="0d0f1a")
    summary_data = [
        ("EAR Labs Scraper — Lead Report", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Powered by WeOps", ""),
        ("", ""),
        ("Total Leads", len(businesses)),
        ("Analyzed", sum(1 for b in businesses if b.get("is_analyzed"))),
        ("With Owner Contact", sum(1 for b in businesses if owners.get(b["id"]))),
        ("Won", sum(1 for b in businesses if b.get("status") == "won")),
        ("Avg Opportunity Score", round(
            sum(analyses[b["id"]].get("opportunity_score", 0) for b in businesses if analyses.get(b["id"])) /
            max(sum(1 for b in businesses if analyses.get(b["id"])), 1), 1
        )),
    ]
    for r, (label, value) in enumerate(summary_data, 1):
        c1 = ws2.cell(row=r, column=1, value=label)
        c2 = ws2.cell(row=r, column=2, value=value)
        for c in [c1, c2]:
            c.fill = bg
            c.font = Font(color="e2e8f0" if r > 1 else "4f6ef7", bold=(r == 1), size=10 if r > 1 else 13)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 20

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
